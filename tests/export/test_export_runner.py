"""Tests for the export-stage runner (``export.runner.run_export_pipeline``).

Verifies the wired runner produces a valid :class:`ExportResult` (processed-data TSV for the
harmonize layer + per-column unique-list workbooks), creates the export directories, honors the
optional ``raw`` layer, and passes the export-paths contract.
"""

from __future__ import annotations

import polars as pl
from openpyxl import load_workbook

from whep_digitize.contracts import (
    LayerDiagnostics,
    PostproDiagnostics,
    PostproResult,
    assert_export_paths_contract,
)
from whep_digitize.export.runner import run_export_pipeline
from whep_digitize.general.config import Config


def _postpro_result() -> PostproResult:
    layer = pl.DataFrame(
        {
            "continent": ["europe", "asia"],
            "polity": ["spain", "japan"],
            "year": ["2000", "2001"],
            "value": ["1", "2"],
        }
    )
    harmonize = layer.with_columns(pl.col("value").cast(pl.Float64))
    diag = LayerDiagnostics(matched_count=0, unmatched_count=0, status="ok")
    return PostproResult(
        harmonize=harmonize,
        clean=layer,
        normalize=layer,
        diagnostics=PostproDiagnostics(clean=diag, standardize_units=diag, harmonize=diag),
    )


def test_run_export_pipeline_returns_valid_result(config: Config) -> None:
    result = run_export_pipeline(config, _postpro_result())

    assert_export_paths_contract(result)  # does not raise
    assert list(result.processed_paths) == ["whep_data_harmonize"]
    assert result.processed_paths["whep_data_harmonize"].is_file()
    # Default lists_to_export keeps only the categorical columns present (continent, polity).
    assert set(result.lists_paths) == {"continent", "polity"}
    assert all(path.is_file() for path in result.lists_paths.values())


def test_run_export_pipeline_creates_export_directories(config: Config) -> None:
    # The runner is responsible for creating the export dirs (they do not exist yet).
    assert not config.paths.data.export.processed.exists()
    run_export_pipeline(config, _postpro_result())
    assert config.paths.data.export.processed.is_dir()
    assert config.paths.data.export.lists.is_dir()


def test_run_export_pipeline_includes_raw_layer(config: Config) -> None:
    raw = pl.DataFrame(
        {
            "continent": ["Europe", "Asia"],
            "polity": ["Spain", "Japan"],
            "year": ["2000", "2001"],
            "value": ["1", "2"],
        }
    )
    result = run_export_pipeline(config, _postpro_result(), raw=raw)

    # With an un-normalized raw layer, the raw values differ, so "raw" is its own sheet.
    workbook = load_workbook(result.lists_paths["polity"], read_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    assert "raw" in sheet_names[0]
    assert sheet_names == ["raw", "clean_normalize_harmonize"]
