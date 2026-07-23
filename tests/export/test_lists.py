"""Tests for the unique-list export (``export.lists``).

Mirrors the R suites ``tests/3-export_pipeline/test-export-lists.R`` and
``tests/testthat/scripts/test_export_column_centric_lists.R``: sheet order + inference, unique
values (blank/NA + radix sort + numeric), layer-by-sheet grouping, union columns, column
resolution, identical-layer merging, path naming, the workbook write, and ``export_lists``
(harmonize-by-default filtering, exclusions, and the filename-collision guard).
"""

from __future__ import annotations

import dataclasses
from pathlib import Path

import polars as pl
import pytest
from openpyxl import load_workbook

from whep_digitize.export.lists.merge import (
    resolve_list_sheet_payloads,
    resolve_lists_export_columns,
)
from whep_digitize.export.lists.unique_values import (
    LISTS_SHEET_ORDER,
    build_column_lists_export_path,
    build_layer_tables_by_sheet,
    collect_union_columns,
    compute_unique_column_values,
    infer_layer_sheet_name,
)
from whep_digitize.export.lists.write import (
    build_column_unique_cache,
    export_lists,
    write_column_lists_workbook,
)
from whep_digitize.general.config import Config
from whep_digitize.general.errors import ValidationError
from whep_digitize.general.options import RuntimeOptions


def _configure_lists(config: Config, columns: tuple[str, ...]) -> Config:
    """Return a copy of ``config`` with ``lists_to_export`` set and the lists dir created."""
    wide = dataclasses.replace(
        config, export_config=dataclasses.replace(config.export_config, lists_to_export=columns)
    )
    wide.paths.data.export.lists.mkdir(parents=True, exist_ok=True)
    return wide


def _read_layout(path: Path) -> dict[str, list[str | None]]:
    """Read a workbook back as ``{sheet_name: column-A values}`` (no header assumed)."""
    workbook = load_workbook(path, read_only=True)
    layout = {ws.title: [row[0].value for row in ws.iter_rows()] for ws in workbook}
    workbook.close()
    return layout


# --------------------------------------------------------------------------- sheet order / infer


def test_lists_sheet_order() -> None:
    assert LISTS_SHEET_ORDER == ("raw", "clean", "normalize", "harmonize")


@pytest.mark.parametrize(
    ("name", "expected"),
    [
        ("whep_data_raw", "raw"),
        ("whep_data_clean", "clean"),
        ("x_normalize", "normalize"),
        ("demo_harmonize", "harmonize"),
    ],
)
def test_infer_layer_sheet_name(name: str, expected: str) -> None:
    assert infer_layer_sheet_name(name) == expected


def test_infer_layer_sheet_name_unknown_raises() -> None:
    with pytest.raises(ValidationError, match="infer layer sheet name"):
        infer_layer_sheet_name("whep_data_standardize")


# --------------------------------------------------------------------------- unique values


def test_compute_unique_prepends_blank_when_na() -> None:
    frame = pl.DataFrame({"polity": ["a", None, "b", None]})
    assert compute_unique_column_values(frame, "polity") == ["(blank)", "a", "b"]


def test_compute_unique_numeric_sort_with_blank_first() -> None:
    frame = pl.DataFrame({"value": [10.0, None, 2.0, 1.0, None]}, schema={"value": pl.Float64})
    assert compute_unique_column_values(frame, "value") == ["(blank)", "1", "2", "10"]


def test_compute_unique_radix_order_uppercase_before_lowercase() -> None:
    frame = pl.DataFrame({"commodity": ["Zebra", "apple", "Banana", "apple"]})
    assert compute_unique_column_values(frame, "commodity") == ["Banana", "Zebra", "apple"]


def test_compute_unique_missing_column_is_empty() -> None:
    assert compute_unique_column_values(pl.DataFrame({"a": ["x"]}), "polity") == []


# --------------------------------------------------------------------------- layer-by-sheet


def test_build_layer_tables_by_sheet_fixed_keys() -> None:
    result = build_layer_tables_by_sheet(
        {
            "whep_data_raw": pl.DataFrame({"polity": ["a", "b"]}),
            "whep_data_harmonize": pl.DataFrame({"polity": ["a", "c"]}),
        }
    )
    assert list(result) == ["raw", "clean", "normalize", "harmonize"]
    assert result["clean"].height == 0
    assert result["normalize"].height == 0


def test_build_layer_tables_by_sheet_empty() -> None:
    result = build_layer_tables_by_sheet({})
    assert list(result) == ["raw", "clean", "normalize", "harmonize"]
    assert all(frame.height == 0 for frame in result.values())


def test_build_layer_tables_by_sheet_unions_multiple_objects() -> None:
    result = build_layer_tables_by_sheet(
        {
            "alpha_raw": pl.DataFrame({"polity": ["a", "b"]}),
            "beta_raw": pl.DataFrame({"polity": ["c"]}),
        }
    )
    assert sorted(result["raw"]["polity"].to_list()) == ["a", "b", "c"]
    assert result["raw"].height == 3


# --------------------------------------------------------------------------- union columns


def test_collect_union_columns_sorted_unique() -> None:
    layer_by_sheet = {
        "raw": pl.DataFrame({"polity": ["a"], "year": ["2020"]}),
        "clean": pl.DataFrame({"polity": ["a"]}),
        "normalize": pl.DataFrame({"unit": ["kg"]}),
        "harmonize": pl.DataFrame({"polity": ["a"], "value": ["1"]}),
    }
    assert collect_union_columns(layer_by_sheet) == ["polity", "unit", "value", "year"]


def test_collect_union_columns_empty() -> None:
    assert collect_union_columns({sheet: pl.DataFrame() for sheet in LISTS_SHEET_ORDER}) == []


# --------------------------------------------------------------------------- cache


def test_build_column_unique_cache_empty_for_missing_columns() -> None:
    layer_by_sheet = {
        "raw": pl.DataFrame({"polity": ["a", "b"]}),
        "clean": pl.DataFrame(),
        "normalize": pl.DataFrame({"polity": ["a", "b"]}),
        "harmonize": pl.DataFrame({"polity": ["a", "c"]}),
    }
    cache = build_column_unique_cache(layer_by_sheet, ["polity"])
    assert cache["clean"]["polity"] == []
    assert cache["raw"]["polity"] == ["a", "b"]
    assert cache["harmonize"]["polity"] == ["a", "c"]


# --------------------------------------------------------------------------- merge grouping


@pytest.mark.parametrize(
    ("values", "expected_sheets"),
    [
        (
            {
                "raw": ["a", "b"],
                "clean": ["a", "b"],
                "normalize": ["a", "b"],
                "harmonize": ["a", "b"],
            },
            ["raw_clean_normalize_harmonize"],
        ),
        (
            {"raw": ["a"], "clean": ["b"], "normalize": ["b"], "harmonize": ["b"]},
            ["raw", "clean_normalize_harmonize"],
        ),
        (
            {"raw": ["a"], "clean": ["b"], "normalize": ["c"], "harmonize": ["c"]},
            ["raw", "clean", "normalize_harmonize"],
        ),
        (
            {"raw": ["a"], "clean": ["b"], "normalize": ["c"], "harmonize": ["d"]},
            ["raw", "clean", "normalize", "harmonize"],
        ),
    ],
)
def test_resolve_list_sheet_payloads(
    values: dict[str, list[str]], expected_sheets: list[str]
) -> None:
    assert list(resolve_list_sheet_payloads(values)) == expected_sheets


# --------------------------------------------------------------------------- column resolution


def test_resolve_lists_export_columns_config_order(config: Config) -> None:
    wide = dataclasses.replace(
        config,
        export_config=dataclasses.replace(config.export_config, lists_to_export=("unit", "polity")),
    )
    assert resolve_lists_export_columns(wide, ["polity", "unit", "value"]) == ["unit", "polity"]


def test_resolve_lists_export_columns_none_present_raises(config: Config) -> None:
    wide = dataclasses.replace(
        config,
        export_config=dataclasses.replace(config.export_config, lists_to_export=("missing",)),
    )
    with pytest.raises(ValidationError, match="none of the configured columns"):
        resolve_lists_export_columns(wide, ["polity", "unit"])


def test_resolve_lists_export_columns_empty_config_raises(config: Config) -> None:
    wide = dataclasses.replace(
        config, export_config=dataclasses.replace(config.export_config, lists_to_export=())
    )
    with pytest.raises(ValidationError, match="must be defined"):
        resolve_lists_export_columns(wide, ["polity"])


def test_resolve_lists_export_columns_duplicate_config_raises(config: Config) -> None:
    wide = dataclasses.replace(
        config,
        export_config=dataclasses.replace(
            config.export_config, lists_to_export=("polity", "polity")
        ),
    )
    with pytest.raises(ValidationError, match="must be unique"):
        resolve_lists_export_columns(wide, ["polity"])


# --------------------------------------------------------------------------- path build


def test_build_column_lists_export_path(config: Config) -> None:
    path = build_column_lists_export_path(config, "polity")
    assert path.name == "unique_polity.xlsx"
    assert path.parent == config.paths.data.export.lists


def test_build_column_lists_export_path_normalizes(config: Config) -> None:
    assert build_column_lists_export_path(config, "Food Balance").name == "unique_food_balance.xlsx"


def test_build_column_lists_export_path_empty_raises(config: Config) -> None:
    with pytest.raises(ValidationError, match="non-empty"):
        build_column_lists_export_path(config, "")


# --------------------------------------------------------------------------- workbook write


def test_write_column_lists_workbook_merges_all_equal(config: Config) -> None:
    wide = _configure_lists(config, ("polity",))
    cache = {layer: {"polity": ["a", "b"]} for layer in LISTS_SHEET_ORDER}
    path = write_column_lists_workbook("polity", cache, wide)
    assert list(_read_layout(path)) == ["raw_clean_normalize_harmonize"]


def test_write_column_lists_workbook_partial_merge(config: Config) -> None:
    wide = _configure_lists(config, ("polity",))
    cache = {
        "raw": {"polity": ["a", "b"]},
        "clean": {"polity": ["c", "d"]},
        "normalize": {"polity": ["e", "f"]},
        "harmonize": {"polity": ["e", "f"]},
    }
    layout = _read_layout(write_column_lists_workbook("polity", cache, wide))
    assert list(layout) == ["raw", "clean", "normalize_harmonize"]
    assert layout["normalize_harmonize"] == ["e", "f"]


def test_write_column_lists_workbook_overwrite_guard(config: Config) -> None:
    wide = _configure_lists(config, ("polity",))
    cache = {layer: {"polity": ["a"]} for layer in LISTS_SHEET_ORDER}
    write_column_lists_workbook("polity", cache, wide)
    with pytest.raises(ValidationError, match="overwrite"):
        write_column_lists_workbook("polity", cache, wide, overwrite=False)


# --------------------------------------------------------------------------- export_lists


def _demo_objects() -> dict[str, pl.DataFrame]:
    frame = pl.DataFrame({"polity": ["a", "b"], "value": ["1", "2"], "year": ["2020", "2021"]})
    return {f"demo_{layer}": frame for layer in LISTS_SHEET_ORDER}


def test_export_lists_honors_configured_columns(config: Config) -> None:
    wide = _configure_lists(config, ("polity",))
    paths = export_lists(wide, _demo_objects())
    assert list(paths) == ["polity"]
    assert "value" not in paths
    assert "year" not in paths
    assert paths["polity"].is_file()


def test_export_lists_no_columns_raises(config: Config) -> None:
    wide = _configure_lists(config, ("polity",))
    empty_objects = {f"demo_{layer}": pl.DataFrame() for layer in LISTS_SHEET_ORDER}
    with pytest.raises(ValidationError, match="no columns found"):
        export_lists(wide, empty_objects)


def test_export_lists_filename_collision_raises(config: Config) -> None:
    wide = _configure_lists(config, ("polity", "Polity"))
    objects = {
        f"demo_{layer}": pl.DataFrame({"polity": ["a"], "Polity": ["a"]})
        for layer in LISTS_SHEET_ORDER
    }
    with pytest.raises(ValidationError, match="same workbook filename"):
        export_lists(wide, objects)


def _export_to_dir(
    config: Config,
    columns: tuple[str, ...],
    objects: dict[str, pl.DataFrame],
    lists_dir: Path,
    *,
    options: RuntimeOptions | None = None,
) -> dict[str, Path]:
    """Run ``export_lists`` writing into ``lists_dir`` (isolates repeated runs for comparison)."""
    lists_dir.mkdir(parents=True, exist_ok=True)
    export = dataclasses.replace(config.paths.data.export, lists=lists_dir)
    data = dataclasses.replace(config.paths.data, export=export)
    cfg = dataclasses.replace(
        config,
        paths=dataclasses.replace(config.paths, data=data),
        export_config=dataclasses.replace(config.export_config, lists_to_export=columns),
    )
    return export_lists(cfg, objects, options=options)


def test_export_lists_bytes_reproducible(config: Config, tmp_path: Path) -> None:
    # Same inputs -> byte-identical workbooks (the pinned created-date defeats xlsxwriter's clock).
    objects = _demo_objects()
    first = _export_to_dir(config, ("polity", "year"), objects, tmp_path / "run1")
    second = _export_to_dir(config, ("polity", "year"), objects, tmp_path / "run2")
    assert list(first) == list(second)
    for column, path in first.items():
        assert path.read_bytes() == second[column].read_bytes()


def test_export_lists_parallel_matches_sequential(config: Config, tmp_path: Path) -> None:
    # ProcessPoolExecutor writes must be byte-identical to sequential (deterministic order + files).
    objects = _demo_objects()
    columns = ("polity", "year")
    sequential = _export_to_dir(
        config,
        columns,
        objects,
        tmp_path / "seq",
        options=RuntimeOptions(export_parallel_workers=1),
    )
    parallel = _export_to_dir(
        config,
        columns,
        objects,
        tmp_path / "par",
        options=RuntimeOptions(export_parallel_workers=2),
    )
    assert list(sequential) == list(parallel)
    for column, path in sequential.items():
        assert path.read_bytes() == parallel[column].read_bytes()
