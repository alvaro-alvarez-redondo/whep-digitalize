"""Parity test: the unique-list export must reproduce R's ``31-lists`` logic + sheet layout.

Builds the four layer frames from the frozen fixture and asserts the ports in
:mod:`whep_digitize.export.lists` reproduce the R goldens: per-(layer, column) unique values
(drop-null, code-point/radix sort, ``"(blank)"`` prepended when any value is missing — including
the un-normalized ``raw`` layer, parity risk #7), the sorted union of columns, the
configured-column resolution, and the identical-layer merge grouping + fixed sheet order per
column. The last check runs the real ``export_lists`` and reads the written workbooks back
(openpyxl — xlsx bytes cannot match across writers, but the sheet names and cell values are the
library-independent logical layout).

If a golden is absent (fresh checkout — goldens are gitignored), the test skips with the
regeneration command rather than failing.
"""

from __future__ import annotations

import dataclasses
import json

import polars as pl
import pytest
from openpyxl import load_workbook
from r_harness import FIXTURES_DIR
from registry import CAPTURES

from whep_digitize.export.lists.merge import resolve_lists_export_columns
from whep_digitize.export.lists.unique_values import (
    build_layer_tables_by_sheet,
    collect_union_columns,
    compute_unique_column_values,
)
from whep_digitize.export.lists.write import export_lists
from whep_digitize.export.processed_data.layers import collect_layer_tables_for_export
from whep_digitize.general.config import Config

_SPEC = CAPTURES["export_lists"]
_FIXTURE_NAME = _SPEC.fixture
assert _FIXTURE_NAME is not None  # this spec always declares a JSON fixture
_FIXTURE_PATH = FIXTURES_DIR / _FIXTURE_NAME

_LAYERS = ("raw", "clean", "normalize", "harmonize")
_OBJECT_NAME = {layer: f"whep_data_{layer}" for layer in _LAYERS}
_COLUMNS = ("continent", "polity", "commodity", "unit")
_FRAME_COLUMNS = ("continent", "polity", "commodity", "unit", "year", "value")
# Matches the R capture's config (a configured-but-absent 'footnotes' must be dropped).
_LISTS_TO_EXPORT = ("continent", "polity", "commodity", "unit", "footnotes")


def _gold(name: str) -> list[str]:
    path = _SPEC.golden_paths()[name]
    if not path.is_file():
        pytest.skip(
            f"Golden {path} missing; regenerate with "
            f"`python tests/parity/capture.py {_SPEC.module}`"
        )
    data: list[str] = json.loads(path.read_text(encoding="utf-8"))
    return data


@pytest.fixture(scope="module")
def layer_frames() -> dict[str, pl.DataFrame]:
    """The fixture as one all-``String`` frame per layer."""
    records = json.loads(_FIXTURE_PATH.read_text(encoding="utf-8"))
    return {
        layer: pl.DataFrame(
            {
                column: pl.Series(records[layer][column], dtype=pl.String)
                for column in _FRAME_COLUMNS
            }
        )
        for layer in _LAYERS
    }


@pytest.fixture(scope="module")
def data_objects(layer_frames: dict[str, pl.DataFrame]) -> dict[str, pl.DataFrame]:
    """The layer frames keyed by their canonical object names."""
    return {_OBJECT_NAME[layer]: layer_frames[layer] for layer in _LAYERS}


@pytest.mark.parity
@pytest.mark.parametrize("layer", _LAYERS)
@pytest.mark.parametrize("column", _COLUMNS)
def test_compute_unique_column_values_parity(
    layer: str, column: str, layer_frames: dict[str, pl.DataFrame]
) -> None:
    result = compute_unique_column_values(layer_frames[layer], column)
    assert result == _gold(f"uniq_{layer}_{column}")


@pytest.mark.parity
def test_union_columns_parity(data_objects: dict[str, pl.DataFrame]) -> None:
    layer_by_sheet = build_layer_tables_by_sheet(collect_layer_tables_for_export(data_objects))
    assert collect_union_columns(layer_by_sheet) == _gold("union_columns")


@pytest.mark.parity
def test_export_columns_parity(config: Config, data_objects: dict[str, pl.DataFrame]) -> None:
    wide = dataclasses.replace(
        config,
        export_config=dataclasses.replace(config.export_config, lists_to_export=_LISTS_TO_EXPORT),
    )
    layer_by_sheet = build_layer_tables_by_sheet(collect_layer_tables_for_export(data_objects))
    union = collect_union_columns(layer_by_sheet)
    assert resolve_lists_export_columns(wide, union) == _gold("export_columns")


@pytest.mark.parity
def test_export_lists_workbook_layout_parity(
    config: Config, data_objects: dict[str, pl.DataFrame]
) -> None:
    wide = dataclasses.replace(
        config,
        export_config=dataclasses.replace(config.export_config, lists_to_export=_LISTS_TO_EXPORT),
    )
    wide.paths.data.export.lists.mkdir(parents=True, exist_ok=True)

    paths = export_lists(wide, data_objects)
    assert sorted(paths) == sorted(_COLUMNS)

    for column in _COLUMNS:
        expected_sheets = _gold(f"sheets_{column}")
        workbook = load_workbook(paths[column], read_only=True)
        assert workbook.sheetnames == expected_sheets
        for sheet_name in expected_sheets:
            worksheet = workbook[sheet_name]
            written = [row[0].value for row in worksheet.iter_rows()]
            # Each sheet's values equal the unique values of any layer in its (merged) name.
            representative_layer = sheet_name.split("_")[0]
            assert written == _gold(f"uniq_{representative_layer}_{column}")
        workbook.close()
