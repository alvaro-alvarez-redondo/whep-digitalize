"""Unit tests for the post-processing data-audit module.

Ports of ``20-audit-config.R`` / ``20-audit-validation.R`` / ``20-audit-export.R`` /
``20-audit-orchestration.R`` (:mod:`whep_digitize.postpro.audit`). Byte parity vs R lives in
``tests/parity/test_data_audit_parity.py``; these pin the behavioral contract without needing R,
including the two preserved quirks (invalid rows retained; the audit regex is stricter than the
float parser, so ``-3.5`` is flagged yet parses).
"""

from __future__ import annotations

from pathlib import Path

import polars as pl
import pytest
from openpyxl import load_workbook

from whep_digitize.general.config import Config
from whep_digitize.general.errors import ValidationError
from whep_digitize.postpro.audit.audit import AuditResult, audit_data_output
from whep_digitize.postpro.audit.config import (
    AUDIT_FINDINGS_COLUMNS,
    NUMERIC_STRING_MESSAGE,
    empty_audit_findings,
    prepare_audit_root,
    resolve_audit_output_paths,
    validate_audit_config,
)
from whep_digitize.postpro.audit.export import export_validation_audit_report
from whep_digitize.postpro.audit.validation import (
    audit_character_non_empty,
    audit_numeric_string,
    build_audit_validation_plan,
    resolve_audit_columns_by_type,
    run_master_validation,
)

_DOCUMENT_VALUE_MAP = {
    "character_non_empty": ("document",),
    "numeric_string": ("value",),
}


def _series(values: list[str | None]) -> pl.Series:
    return pl.Series(values, dtype=pl.String)


# --------------------------------------------------------------------------- config / findings


def test_empty_audit_findings_schema() -> None:
    findings = empty_audit_findings()
    assert findings.height == 0
    assert tuple(findings.columns) == AUDIT_FINDINGS_COLUMNS
    assert findings.schema["row_index"] == pl.Int64
    assert findings.schema["audit_column"] == pl.String


def test_validate_audit_config_accepts_default(config: Config) -> None:
    validate_audit_config(config)  # does not raise


def test_resolve_audit_output_paths_joins_name() -> None:
    path = resolve_audit_output_paths(Path("data") / "audit", "whep_audit.xlsx")
    assert path == Path("data") / "audit" / "whep_audit.xlsx"


def test_resolve_audit_output_paths_rejects_blank_name() -> None:
    with pytest.raises(ValidationError):
        resolve_audit_output_paths(Path("data"), "")


def test_prepare_audit_root_deletes_existing(tmp_path: Path) -> None:
    audit_dir = tmp_path / "audit"
    audit_dir.mkdir()
    (audit_dir / "stale.xlsx").write_text("x", encoding="utf-8")
    assert prepare_audit_root(audit_dir) is True
    assert not audit_dir.exists()


def test_prepare_audit_root_missing_is_false(tmp_path: Path) -> None:
    assert prepare_audit_root(tmp_path / "does_not_exist") is False


# --------------------------------------------------------------------------- validators


def test_character_non_empty_flags_null_blank_and_whitespace() -> None:
    dataset = pl.DataFrame({"document": _series(["a.xlsx", "", "   ", None, "b.xlsx"])})
    findings = audit_character_non_empty(dataset, "document")
    assert findings.get_column("row_index").to_list() == [2, 3, 4]
    assert findings.get_column("audit_type").unique().to_list() == ["character_non_empty"]
    assert findings.get_column("audit_column").unique().to_list() == ["document"]


def test_character_non_empty_all_valid_is_empty() -> None:
    dataset = pl.DataFrame({"document": _series(["a.xlsx", "b.xlsx"])})
    assert audit_character_non_empty(dataset, "document").height == 0


def test_character_non_empty_missing_column_raises() -> None:
    with pytest.raises(ValidationError):
        audit_character_non_empty(pl.DataFrame({"other": _series(["x"])}), "document")


def test_numeric_string_flags_non_numeric_but_skips_null() -> None:
    # "-3.5" (negative), "3." (trailing dot), ".5" (no leading digit), "1e5", "+3" all flagged;
    # null is skipped; plain ints/decimals pass.
    dataset = pl.DataFrame(
        {"value": _series(["10", "20.5", "bad", "-3.5", None, "3.", ".5", "1e5", "+3", "007"])}
    )
    findings = audit_numeric_string(dataset, "value")
    assert findings.get_column("row_index").to_list() == [3, 4, 6, 7, 8, 9]
    assert findings.get_column("audit_type").unique().to_list() == ["numeric_string"]


def test_numeric_string_defaults_to_value_column() -> None:
    dataset = pl.DataFrame({"value": _series(["10", "bad"])})
    assert audit_numeric_string(dataset).get_column("row_index").to_list() == [2]


# --------------------------------------------------------------------------- validation plan


def test_build_plan_expands_type_and_columns_in_order() -> None:
    plan = build_audit_validation_plan(
        {"character_non_empty": ("a", "b"), "numeric_string": ("value",)},
        ("character_non_empty", "numeric_string"),
    )
    assert plan.get_column("audit_type").to_list() == [
        "character_non_empty",
        "character_non_empty",
        "numeric_string",
    ]
    assert plan.get_column("column_name").to_list() == ["a", "b", "value"]


def test_build_plan_rejects_empty_column_vector() -> None:
    with pytest.raises(ValidationError):
        build_audit_validation_plan({"numeric_string": ()}, ("numeric_string",))


# --------------------------------------------------------------------------- master validation


def _document_value_dataset() -> pl.DataFrame:
    return pl.DataFrame(
        {
            "document": _series(["a.xlsx", "", "b.xlsx"]),
            "value": _series(["10", "bad", "-3.5"]),
        }
    )


def test_master_validation_orders_findings_by_plan() -> None:
    result = run_master_validation(_document_value_dataset(), _DOCUMENT_VALUE_MAP)
    # character_non_empty (document) first, then numeric_string (value).
    assert result.findings.get_column("audit_column").to_list() == ["document", "value", "value"]
    assert result.findings.get_column("row_index").to_list() == [2, 2, 3]
    # invalid_row_index is sorted + unique across both validators.
    assert result.invalid_row_index == (2, 3)


def test_master_validation_warns_and_skips_unsupported_types() -> None:
    with pytest.warns(UserWarning, match="unsupported audit types"):
        result = run_master_validation(
            _document_value_dataset(),
            {"numeric_string": ("value",), "unknown_check": ("document",)},
        )
    assert result.findings.get_column("audit_type").unique().to_list() == ["numeric_string"]


def test_master_validation_selected_validations_filters() -> None:
    result = run_master_validation(
        _document_value_dataset(), _DOCUMENT_VALUE_MAP, selected_validations=["numeric_string"]
    )
    # Only numeric_string runs: "bad" (row 2) and "-3.5" (row 3) are flagged; document is skipped.
    assert result.findings.get_column("audit_column").unique().to_list() == ["value"]
    assert result.invalid_row_index == (2, 3)


def test_master_validation_no_supported_returns_empty() -> None:
    with pytest.warns(UserWarning, match="unsupported audit types"):
        result = run_master_validation(_document_value_dataset(), {"unknown_check": ("document",)})
    assert result.findings.height == 0
    assert result.invalid_row_index == ()


def test_master_validation_clean_dataset_is_empty() -> None:
    clean = pl.DataFrame({"document": _series(["a.xlsx"]), "value": _series(["10"])})
    result = run_master_validation(clean, _DOCUMENT_VALUE_MAP)
    assert result.findings.height == 0
    assert result.invalid_row_index == ()


# --------------------------------------------------------------------------- column resolution


def test_resolve_columns_default_from_config(config: Config) -> None:
    resolved = resolve_audit_columns_by_type(config)
    assert resolved["character_non_empty"] == tuple(dict.fromkeys(config.audit_columns))
    assert resolved["numeric_string"] == ("value",)


def test_resolve_columns_override_wins(config: Config) -> None:
    override = {"character_non_empty": ["document"], "numeric_string": ["value"]}
    assert resolve_audit_columns_by_type(config, override) == {
        "character_non_empty": ("document",),
        "numeric_string": ("value",),
    }


# --------------------------------------------------------------------------- export (openpyxl)


def _small_findings(row_index: int, audit_column: str) -> pl.DataFrame:
    return pl.DataFrame(
        {
            "row_index": pl.Series([row_index], dtype=pl.Int64),
            "audit_column": _series([audit_column]),
            "audit_type": _series(["numeric_string"]),
            "audit_message": _series([NUMERIC_STRING_MESSAGE]),
        }
    )


def test_export_returns_none_when_empty(config: Config, tmp_path: Path) -> None:
    empty = pl.DataFrame({"document": _series([]), "value": _series([])})
    output_path = tmp_path / "audit.xlsx"
    result = export_validation_audit_report(empty, config, empty_audit_findings(), output_path)
    assert result is None
    assert not output_path.exists()


def test_export_highlights_flagged_cell_with_document_sort(config: Config, tmp_path: Path) -> None:
    # Two invalid rows; findings flag the "bad" value. source_row_index 1 = b.xlsx row.
    audit_dt = pl.DataFrame(
        {"document": _series(["b.xlsx", "a.xlsx"]), "value": _series(["bad", "10"])}
    )
    output_path = tmp_path / "audit.xlsx"
    result = export_validation_audit_report(
        audit_dt, config, _small_findings(1, "value"), output_path
    )
    assert result == output_path

    worksheet = load_workbook(output_path).active
    assert worksheet.title == "audit_report"
    # Header, then rows sorted by document (a.xlsx before b.xlsx).
    assert [worksheet.cell(1, col).value for col in (1, 2)] == ["document", "value"]
    assert worksheet.cell(2, 1).value == "a.xlsx"
    assert worksheet.cell(3, 1).value == "b.xlsx"
    # The flagged "bad" cell (row 3, col 2) is highlighted; the valid "10" cell is not.
    flagged = worksheet.cell(3, 2)
    assert flagged.value == "bad"
    assert flagged.fill.fgColor.rgb == "FFFFB84D"
    assert flagged.font.bold is True
    assert flagged.border.left.style == "thick"
    assert worksheet.cell(2, 2).fill.fill_type is None


def test_export_writes_note_when_no_rows_but_findings(config: Config, tmp_path: Path) -> None:
    empty = pl.DataFrame({"document": _series([]), "value": _series([])})
    output_path = tmp_path / "audit.xlsx"
    result = export_validation_audit_report(empty, config, _small_findings(1, "value"), output_path)
    assert result == output_path
    worksheet = load_workbook(output_path).active
    assert worksheet.cell(1, 1).value == "note"
    assert worksheet.cell(2, 1).value == "No audit findings detected for this dataset."


# --------------------------------------------------------------------------- orchestration


def _full_dataset() -> pl.DataFrame:
    # Row 2 has a blank document (character_non_empty) and "-3.5" value (numeric_string).
    return pl.DataFrame(
        {
            "continent": _series(["Asia", "Europe", "Africa"]),
            "polity": _series(["Japan", "France", "Egypt"]),
            "commodity": _series(["wheat", "wheat", "wheat"]),
            "variable": _series(["production", "production", "production"]),
            "unit": _series(["tonnes", "tonnes", "tonnes"]),
            "yearbook": _series(["yb", "yb", "yb"]),
            "document": _series(["a.xlsx", "", "c.xlsx"]),
            "value": _series(["10", "-3.5", "bad"]),
        }
    )


def test_audit_data_output_keeps_invalid_rows_and_parses_value(config: Config) -> None:
    result = audit_data_output(_full_dataset(), config)
    assert isinstance(result, AuditResult)
    # All rows retained (invalid rows are NOT dropped).
    assert result.audited.height == 3
    # The divergence: "-3.5" is flagged yet still parses to -3.5; "bad" -> null.
    assert result.audited.schema["value"] == pl.Float64
    assert result.audited.get_column("value").to_list() == [10.0, -3.5, None]


def test_audit_data_output_findings_capture_divergence(config: Config) -> None:
    result = audit_data_output(_full_dataset(), config)
    flagged = result.findings.filter(
        (pl.col("audit_column") == "value") & (pl.col("audit_type") == "numeric_string")
    )
    # Row 2 ("-3.5") and row 3 ("bad") are both flagged by the stricter regex.
    assert flagged.get_column("row_index").to_list() == [2, 3]
    assert result.invalid_row_index == (2, 3)


def test_audit_data_output_writes_report_when_findings(config: Config) -> None:
    result = audit_data_output(_full_dataset(), config)
    assert result.report_path is not None
    assert result.report_path.exists()
    assert result.report_path == config.paths.data.audit.audit_file_path


def test_audit_data_output_skips_report_when_clean(config: Config) -> None:
    clean = pl.DataFrame(
        {
            "continent": _series(["Asia"]),
            "polity": _series(["Japan"]),
            "commodity": _series(["wheat"]),
            "variable": _series(["production"]),
            "unit": _series(["tonnes"]),
            "yearbook": _series(["yb"]),
            "document": _series(["a.xlsx"]),
            "value": _series(["10"]),
        }
    )
    result = audit_data_output(clean, config)
    assert result.report_path is None
    assert not config.paths.data.audit.audit_file_path.exists()
    assert result.findings.height == 0
    assert result.audited.get_column("value").to_list() == [10.0]


def test_audit_data_output_without_value_column_leaves_frame(config: Config) -> None:
    dataset = pl.DataFrame({"document": _series(["a.xlsx", "b.xlsx"])})
    override = {"character_non_empty": ["document"]}
    result = audit_data_output(dataset, config, audit_columns_by_type=override)
    assert "value" not in result.audited.columns
    assert result.audited.get_column("document").to_list() == ["a.xlsx", "b.xlsx"]
    assert result.findings.height == 0
