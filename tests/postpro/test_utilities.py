"""Unit tests for the post-processing utilities modules.

Ports of ``21-output-roots.R`` / ``21-diagnostics.R`` / ``21-template-rules.R`` /
``21-runtime-cache.R`` (:mod:`whep_digitize.postpro.utilities`). Byte parity vs R for
``read_rule_table`` + ``build_layer_diagnostics`` lives in
``tests/parity/test_utilities_parity.py``; these pin the behavioral contract without needing R.
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

import polars as pl
import pytest
from openpyxl import Workbook, load_workbook

from whep_digitize.general.config import Config
from whep_digitize.general.errors import ValidationError
from whep_digitize.postpro.utilities.diagnostics import build_layer_diagnostics
from whep_digitize.postpro.utilities.output_roots import (
    PostproOutputPaths,
    get_postpro_output_paths,
    initialize_postpro_output_root,
)
from whep_digitize.postpro.utilities.payload_cache import (
    RuntimeCacheSettings,
    StagePayloadBundle,
    build_stage_payload_cache_key,
    clear_stage_payload_memory_cache,
    get_cached_stage_payload_bundle,
    prune_runtime_cache_entries,
    resolve_stage_runtime_cache_settings,
)
from whep_digitize.postpro.utilities.templates import (
    generate_postpro_rule_templates,
    load_stage_rule_payloads,
    read_rule_table,
    write_stage_rule_template,
)

_CANONICAL_HEADER = (
    "clean_column_source,clean_value_source_raw,clean_value_source,"
    "clean_column_target,clean_value_target_raw,clean_value_target"
)
_RULE_FIXTURE = Path(__file__).parents[1] / "fixtures" / "synthetic" / "clean_rules_sample.xlsx"


@pytest.fixture(autouse=True)
def _reset_cache() -> Iterator[None]:
    clear_stage_payload_memory_cache()
    yield
    clear_stage_payload_memory_cache()


def _write_clean_rule_csv(path: Path, source: str = "commodity") -> None:
    path.write_text(f"{_CANONICAL_HEADER}\n{source},wheat,WHEAT,unit,t,tonne\n", encoding="utf-8")


# --------------------------------------------------------------------------- output_roots


def test_get_postpro_output_paths_from_config(config: Config) -> None:
    paths = get_postpro_output_paths(config)
    assert isinstance(paths, PostproOutputPaths)
    audit = config.paths.data.audit
    assert paths.audit_dir == audit.audit_dir
    assert paths.diagnostics_dir == audit.diagnostics_dir
    assert paths.templates_dir == audit.templates_dir
    assert paths.runtime_cache_dir == audit.runtime_cache_dir


def test_initialize_creates_all_output_dirs(config: Config) -> None:
    paths = initialize_postpro_output_root(config)
    for directory in (
        paths.audit_root_dir,
        paths.audit_dir,
        paths.diagnostics_dir,
        paths.templates_dir,
        paths.runtime_cache_dir,
    ):
        assert directory.is_dir()


# --------------------------------------------------------------------------- diagnostics


def test_build_layer_diagnostics_matched() -> None:
    audit = pl.DataFrame({"affected_rows": pl.Series([2, 3], dtype=pl.Int64)})
    diagnostics = build_layer_diagnostics("clean", 10, 10, audit)
    assert diagnostics.matched_count == 5
    assert diagnostics.unmatched_count == 5
    assert diagnostics.status == "pass"
    assert diagnostics.messages == ("Rules applied successfully",)


def test_build_layer_diagnostics_empty_audit_warns() -> None:
    audit = pl.DataFrame({"affected_rows": pl.Series([], dtype=pl.Int64)})
    diagnostics = build_layer_diagnostics("clean", 5, 5, audit)
    assert diagnostics.matched_count == 0
    assert diagnostics.unmatched_count == 5
    assert diagnostics.status == "warn"
    assert diagnostics.messages == ("No rows matched available rules",)


def test_build_layer_diagnostics_missing_affected_rows_column() -> None:
    # A non-empty audit table without affected_rows sums to 0 (R sum(NULL) == 0).
    audit = pl.DataFrame({"other": pl.Series(["x"], dtype=pl.String)})
    diagnostics = build_layer_diagnostics("clean", 3, 3, audit)
    assert diagnostics.matched_count == 0
    assert diagnostics.unmatched_count == 3


def test_build_layer_diagnostics_rejects_negative_rows() -> None:
    audit = pl.DataFrame({"affected_rows": pl.Series([], dtype=pl.Int64)})
    with pytest.raises(ValidationError):
        build_layer_diagnostics("clean", -1, 0, audit)


def test_build_layer_diagnostics_rejects_blank_layer() -> None:
    audit = pl.DataFrame({"affected_rows": pl.Series([], dtype=pl.Int64)})
    with pytest.raises(ValidationError):
        build_layer_diagnostics("", 1, 1, audit)


# --------------------------------------------------------------------------- read_rule_table


def test_read_rule_table_xlsx_strips_prefix_skips_nonmatching_all_text() -> None:
    rules = read_rule_table(_RULE_FIXTURE)
    assert rules.columns == [
        "column_source",
        "value_source_raw",
        "value_source",
        "column_target",
        "value_target_raw",
        "value_target",
    ]
    assert rules.height == 2  # guidance sheet skipped
    assert rules.schema["value_source_raw"] == pl.String
    # Numeric-looking codes keep their exact source string.
    assert rules.get_column("value_source_raw").to_list() == ["wheat", "007"]
    assert rules.get_column("value_target_raw").to_list() == ["t", "1000.0"]


def test_read_rule_table_csv_all_text(tmp_path: Path) -> None:
    csv_path = tmp_path / "clean_rules.csv"
    csv_path.write_text("code\n007\n1000.0\n", encoding="utf-8")
    rules = read_rule_table(csv_path)
    assert rules.schema["code"] == pl.String
    assert rules.get_column("code").to_list() == ["007", "1000.0"]


def test_read_rule_table_unsupported_extension_raises(tmp_path: Path) -> None:
    bad = tmp_path / "rules.txt"
    bad.write_text("x", encoding="utf-8")
    with pytest.raises(ValidationError):
        read_rule_table(bad)


def test_read_rule_table_missing_file_raises(tmp_path: Path) -> None:
    with pytest.raises(ValidationError):
        read_rule_table(tmp_path / "nope.xlsx")


def test_read_rule_table_no_matching_sheet_raises(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "notes"
    sheet.append(["note", "detail"])  # not the canonical schema
    sheet.append(["a", "b"])
    path = tmp_path / "clean_bad.xlsx"
    workbook.save(path)
    with pytest.raises(ValidationError):
        read_rule_table(path)


# --------------------------------------------------------------------------- templates


def test_write_stage_rule_template_creates_two_sheets(tmp_path: Path) -> None:
    template_path = write_stage_rule_template(tmp_path)
    assert template_path.is_file()
    workbook = load_workbook(template_path)
    assert workbook.sheetnames == ["clean_harmonize_template", "guidance"]
    header = [cell.value for cell in workbook["clean_harmonize_template"][1]]
    assert header == [
        "column_source",
        "value_source_raw",
        "value_source",
        "column_target",
        "value_target_raw",
        "value_target",
    ]
    assert workbook["guidance"]["A1"].value == "note"


def test_write_stage_rule_template_respects_no_overwrite(tmp_path: Path) -> None:
    template_path = write_stage_rule_template(tmp_path)
    template_path.write_text("sentinel", encoding="utf-8")  # corrupt it
    returned = write_stage_rule_template(tmp_path, overwrite=False)
    assert returned == template_path
    assert template_path.read_text(encoding="utf-8") == "sentinel"  # untouched


def test_generate_postpro_rule_templates(config: Config) -> None:
    template_path = generate_postpro_rule_templates(config)
    assert template_path.is_file()
    assert template_path.parent == config.paths.data.audit.templates_dir


def test_load_stage_rule_payloads_discovers_orders_and_filters(config: Config) -> None:
    cleaning = config.paths.data.import_.cleaning
    cleaning.mkdir(parents=True, exist_ok=True)
    _write_clean_rule_csv(cleaning / "clean_b.csv")
    _write_clean_rule_csv(cleaning / "clean_a.csv")
    _write_clean_rule_csv(cleaning / "harmonize_x.csv")  # wrong stage prefix
    (cleaning / "ignore.txt").write_text("x", encoding="utf-8")  # wrong extension

    payloads = load_stage_rule_payloads(config, "clean")
    assert [payload.rule_file_id for payload in payloads] == ["clean_a.csv", "clean_b.csv"]
    assert payloads[0].raw_rules.get_column("clean_column_source").to_list() == ["commodity"]


# --------------------------------------------------------------------------- payload_cache


def test_resolve_runtime_cache_settings_off_by_default(config: Config) -> None:
    settings = resolve_stage_runtime_cache_settings(config)
    assert settings.enabled is False
    assert settings.max_entries >= 1


def test_build_cache_key_no_rule_files(config: Config) -> None:
    assert build_stage_payload_cache_key(config, "clean") == "clean::<no_rule_files>"


def test_build_cache_key_reflects_file_content(config: Config) -> None:
    cleaning = config.paths.data.import_.cleaning
    cleaning.mkdir(parents=True, exist_ok=True)
    rule_file = cleaning / "clean_r.csv"
    _write_clean_rule_csv(rule_file, source="commodity")
    key_first = build_stage_payload_cache_key(config, "clean")
    assert key_first.startswith("clean::clean_r.csv::")
    assert "@@" in key_first

    _write_clean_rule_csv(rule_file, source="polity")  # different md5
    assert build_stage_payload_cache_key(config, "clean") != key_first


def test_prune_keeps_lowest_sorted_keys() -> None:
    entries = {name: StagePayloadBundle(name, ()) for name in ("c", "a", "b")}
    pruned = prune_runtime_cache_entries(entries, 2)
    assert sorted(pruned) == ["a", "b"]
    assert len(entries) == 3  # input not mutated


def test_get_cached_bundle_disabled_builds(config: Config) -> None:
    cleaning = config.paths.data.import_.cleaning
    cleaning.mkdir(parents=True, exist_ok=True)
    _write_clean_rule_csv(cleaning / "clean_r.csv")

    bundle = get_cached_stage_payload_bundle(config, "clean")
    assert bundle.cache_key.startswith("clean::")
    assert [payload.rule_file_id for payload in bundle.canonical_payloads] == ["clean_r.csv"]
    # coerce_rule_schema stripped the clean_ prefix.
    assert "column_source" in bundle.canonical_payloads[0].canonical_rules.columns


def test_get_cached_bundle_enabled_uses_memory_then_disk(config: Config) -> None:
    cleaning = config.paths.data.import_.cleaning
    cleaning.mkdir(parents=True, exist_ok=True)
    _write_clean_rule_csv(cleaning / "clean_r.csv")
    settings = RuntimeCacheSettings(enabled=True, cache_file_name="cache.pkl", max_entries=128)

    first = get_cached_stage_payload_bundle(config, "clean", settings=settings)
    cache_file = config.paths.data.audit.runtime_cache_dir / "cache.pkl"
    assert cache_file.is_file()  # persisted to disk

    # Second call hits the in-memory cache (same object).
    assert get_cached_stage_payload_bundle(config, "clean", settings=settings) is first

    # Clearing memory forces a disk reload: a new object with the same key + payloads.
    clear_stage_payload_memory_cache()
    reloaded = get_cached_stage_payload_bundle(config, "clean", settings=settings)
    assert reloaded is not first
    assert reloaded.cache_key == first.cache_key
    assert [payload.rule_file_id for payload in reloaded.canonical_payloads] == ["clean_r.csv"]
