"""Unit tests for the post-processing diagnostics modules.

Ports of ``25-preflight.R`` / ``25-rule-summaries.R`` / ``25-standardize-summaries.R`` /
``25-diagnostics-output.R`` (:mod:`whep_digitize.postpro.diagnostics`). Byte parity vs R for the
summary tables lives in ``tests/parity/test_diagnostics_parity.py``; these mirror the R testthat
cases without needing R.
"""

from __future__ import annotations

import polars as pl
import pytest
from openpyxl import load_workbook

from whep_digitize.general.config import Config
from whep_digitize.general.errors import WhepError
from whep_digitize.postpro.diagnostics.output import (
    build_last_rule_wins_overwrite_subset,
    build_postpro_diagnostics,
    persist_postpro_audit,
)
from whep_digitize.postpro.diagnostics.preflight import (
    PreflightResult,
    assert_postpro_preflight,
    collect_postpro_preflight,
)
from whep_digitize.postpro.diagnostics.rule_summaries import (
    build_stage_rule_catalog_from_payloads,
    build_unmatched_rule_summary,
    summarize_stage_rules,
)
from whep_digitize.postpro.diagnostics.standardize_summaries import (
    build_standardize_rule_catalog,
    build_unmatched_standardize_rule_summary,
    summarize_standardize_rules,
)
from whep_digitize.postpro.utilities.templates import RulePayload


def _s(values: list[str | None]) -> pl.Series:
    return pl.Series(values, dtype=pl.String)


def _i(values: list[int | None]) -> pl.Series:
    return pl.Series(values, dtype=pl.Int64)


def _f(values: list[float | None]) -> pl.Series:
    return pl.Series(values, dtype=pl.Float64)


def _make_preflight_dirs(config: Config) -> None:
    for directory in (
        config.paths.data.import_.cleaning,
        config.paths.data.import_.harmonization,
        config.paths.data.audit.templates_dir,
        config.paths.data.audit.diagnostics_dir,
    ):
        directory.mkdir(parents=True, exist_ok=True)


# --------------------------------------------------------------------------- preflight


def test_preflight_flags_invalid_naming(config: Config) -> None:
    _make_preflight_dirs(config)
    (config.paths.data.import_.cleaning / "bad_clean_name.xlsx").touch()
    (config.paths.data.import_.harmonization / "bad_harmonize_name.xlsx").touch()
    result = collect_postpro_preflight(config, ["unit", "value", "commodity"])
    assert result.checks["cleaning_pattern_ok"] is False
    assert result.checks["harmonize_pattern_ok"] is False
    assert any("clean stage" in issue for issue in result.issues)
    assert any("harmonize stage" in issue for issue in result.issues)


def test_preflight_detects_missing_columns(config: Config) -> None:
    _make_preflight_dirs(config)
    (config.paths.data.import_.cleaning / "clean_rules.xlsx").touch()
    (config.paths.data.import_.harmonization / "harmonize_rules.xlsx").touch()
    result = collect_postpro_preflight(config, ["unit", "value", "item"])
    assert result.passed is False
    assert any("missing expected columns" in issue for issue in result.issues)


def test_preflight_passes_when_clean(config: Config) -> None:
    _make_preflight_dirs(config)
    (config.paths.data.import_.cleaning / "clean_rules.xlsx").touch()
    (config.paths.data.import_.harmonization / "harmonize_rules.xlsx").touch()
    result = collect_postpro_preflight(config, ["unit", "value", "commodity"])
    assert result.passed is True
    assert result.issues == ()


def test_assert_preflight_aborts() -> None:
    bad = PreflightResult(passed=False, issues=("[clean stage] missing file",), checks={})
    with pytest.raises(WhepError, match="preflight checks failed"):
        assert_postpro_preflight(bad)


# --------------------------------------------------------------------------- rule summaries


def _clean_audit() -> pl.DataFrame:
    return pl.DataFrame(
        {
            "loop": _i([1, 2]),
            "rule_file_identifier": _s(["clean_rules.csv", "clean_rules.csv"]),
            "column_source": _s(["commodity", "commodity"]),
            "value_source_raw": _s(["wheat", "rice"]),
            "value_source_result": _s(["wheat_clean", "rice_clean"]),
            "column_target": _s(["unit", "variable"]),
            "value_target_raw": _s(["kg", "kg"]),
            "value_target_result": _s(["kilogram", "gram"]),
            "affected_rows": _i([5, 3]),
        }
    )


def test_summarize_stage_rules_fills_from_result() -> None:
    summary = summarize_stage_rules(_clean_audit())
    assert summary.columns[:2] == ["loop", "affected_rows"]
    assert "execution_stage" not in summary.columns
    assert summary.get_column("value_source").to_list() == ["wheat_clean", "rice_clean"]
    assert summary.get_column("value_target").to_list() == ["kilogram", "gram"]


def test_build_unmatched_rule_summary_anti_joins() -> None:
    catalog = pl.DataFrame(
        {
            "rule_file_identifier": _s(["clean_rules.csv", "clean_rules.csv"]),
            "column_source": _s(["commodity", "polity"]),
            "value_source_raw": _s(["wheat", "spain"]),
            "value_source": _s(["wheat", "spain"]),
            "column_target": _s(["unit", "continent"]),
            "value_target_raw": _s(["kg", "eu"]),
            "value_target": _s(["kg", "eu"]),
        }
    )
    unmatched = build_unmatched_rule_summary(catalog, summarize_stage_rules(_clean_audit()))
    assert unmatched.height == 1
    assert unmatched.get_column("column_source").to_list() == ["polity"]
    assert unmatched.get_column("affected_rows").to_list() == [0]


def test_build_stage_rule_catalog_from_payloads() -> None:
    payload = RulePayload(
        rule_file_id="clean_rules.xlsx",
        rule_file_path="/rules/clean_rules.xlsx",
        raw_rules=pl.DataFrame(
            {
                "column_source": _s(["commodity"]),
                "value_source_raw": _s(["wheat"]),
                "value_source": _s(["wheat"]),
                "column_target": _s(["unit"]),
                "value_target_raw": _s(["kg"]),
                "value_target": _s(["kg"]),
            }
        ),
    )
    catalog = build_stage_rule_catalog_from_payloads([payload])
    assert catalog.height == 1
    assert catalog.get_column("rule_file_identifier").to_list() == ["clean_rules.xlsx"]


# --------------------------------------------------------------------------- standardize summaries


def test_summarize_standardize_rules() -> None:
    audit = pl.DataFrame(
        {
            "rule_file_identifier": _s(["standardize_units_rules.xlsx"]),
            "commodity_key": _s(["all commodity"]),
            "unit_source": _s(["kg"]),
            "unit_target": _s(["t"]),
            "unit_factor": _f([0.001]),
            "unit_offset": _f([0.0]),
            "affected_rows": _i([3]),
        }
    )
    summary = summarize_standardize_rules(audit)
    assert summary.columns[:2] == ["affected_rows", "rule_file_identifier"]
    assert summary.get_column("commodity_key").to_list() == ["all commodity"]


def test_build_unmatched_standardize_counts_branch() -> None:
    catalog = build_standardize_rule_catalog(
        pl.DataFrame(
            {
                "source_rule_file": _s(["rules.xlsx", "rules.xlsx"]),
                "commodity_key": _s(["all commodity", "wheat"]),
                "unit_source": _s(["kg", "tonne"]),
                "unit_target": _s(["g", "kg"]),
                "unit_factor": _f([1000.0, 1000.0]),
                "unit_offset": _f([0.0, 0.0]),
            }
        )
    )
    matched_summary = pl.DataFrame(
        {
            "affected_rows": _i([2]),
            "rule_file_identifier": _s(["rules.xlsx"]),
            "commodity_key": _s(["corn"]),
            "unit_source": _s(["kg"]),
            "unit_target": _s(["g"]),
            "unit_factor": _f([1000.0]),
            "unit_offset": _f([0.0]),
            "source_unit_raw": _s([None]),
            "detected_prefix": _f([None]),
            "unit_factor_effective": _f([None]),
        }
    )
    counts = pl.DataFrame(
        {
            "rule_commodity_match_key": _s(["all commodity"]),
            "applied_commodity_match_key": _s(["corn"]),
            "unit_source_key": _s(["kg"]),
            "affected_rows": _i([2]),
        }
    )
    unmatched = build_unmatched_standardize_rule_summary(catalog, matched_summary, counts)
    assert unmatched.height == 1
    assert unmatched.get_column("commodity_key").to_list() == ["wheat"]
    assert unmatched.get_column("unit_source").to_list() == ["tonne"]
    assert unmatched.get_column("affected_rows").to_list() == [0]


# --------------------------------------------------------------------------- output


def test_build_postpro_diagnostics_creates_three_summaries() -> None:
    empty_stage = pl.DataFrame(schema={"loop": pl.Int64, "affected_rows": pl.Int64})
    empty_std = pl.DataFrame(schema={"affected_rows": pl.Int64})
    summaries = build_postpro_diagnostics(empty_stage, empty_stage, empty_std)
    assert summaries.clean_rule_summary.height == 0
    assert summaries.harmonize_rule_summary.height == 0
    assert summaries.standardize_rule_summary.height == 0


def test_build_last_rule_wins_overwrite_subset() -> None:
    final = pl.DataFrame({"polity": _s(["polityA", "polityB"]), "notes": _s(["note a", "note b"])})
    events = pl.DataFrame(
        {
            "execution_stage": _s(["harmonize"]),
            "rule_file_identifier": _s(["harmonize_rules.xlsx"]),
            "column_target": _s(["notes"]),
            "row_id": _i([2]),
        }
    )
    subset = build_last_rule_wins_overwrite_subset(final, events)
    assert subset.height == 1
    assert subset.get_column("row_id").to_list() == [2]
    assert subset.get_column("overwrite_event_count").to_list() == [1]
    assert subset.get_column("overwritten_columns").to_list() == ["notes"]
    assert subset.get_column("polity").to_list() == ["polityB"]


def test_persist_postpro_audit_writes_workbooks(config: Config) -> None:
    empty_stage = pl.DataFrame(
        schema={
            "loop": pl.Int64,
            "rule_file_identifier": pl.String,
            "column_source": pl.String,
            "value_source_raw": pl.String,
            "value_source": pl.String,
            "column_target": pl.String,
            "value_target_raw": pl.String,
            "value_target": pl.String,
            "affected_rows": pl.Int64,
        }
    )
    std_audit = pl.DataFrame(
        {
            "rule_file_identifier": _s(["standardize_units_rules.xlsx"]),
            "commodity_key": _s(["wheat"]),
            "unit_source": _s(["kg"]),
            "unit_target": _s(["t"]),
            "unit_factor": _f([0.001]),
            "unit_offset": _f([0.0]),
            "affected_rows": _i([2]),
        }
    )
    std_rules = pl.DataFrame(
        {
            "commodity_key": _s(["wheat"]),
            "unit_source": _s(["kg"]),
            "unit_target": _s(["t"]),
            "source_rule_file": _s(["standardize_units_rules.xlsx"]),
        }
    )
    final = pl.DataFrame({"polity": _s(["polityA", "polityB"]), "notes": _s(["note a", "note b"])})
    events = pl.DataFrame(
        {
            "execution_stage": _s(["harmonize"]),
            "rule_file_identifier": _s(["harmonize_rules.xlsx"]),
            "column_target": _s(["notes"]),
            "row_id": _i([2]),
        }
    )
    paths = persist_postpro_audit(
        empty_stage, empty_stage, std_audit, std_rules, final, events, config
    )

    assert set(paths) == {
        "clean_audit",
        "harmonize_audit",
        "standardize_audit",
        "last_rule_wins_overwrites",
    }
    for name in ("clean_audit", "harmonize_audit", "standardize_audit"):
        assert paths[name].is_file()
        assert load_workbook(paths[name]).sheetnames == ["matched_rules", "unmatched_rules"]
    overwrite = load_workbook(paths["last_rule_wins_overwrites"])
    assert overwrite.sheetnames == ["last_rule_wins_overwrites"]
    standardize = load_workbook(paths["standardize_audit"])["matched_rules"]
    header = [cell.value for cell in standardize[1]]
    assert header[:2] == ["affected_rows", "rule_file_identifier"]
    assert "loop" not in header
