"""Postpro / diagnostics — cross-stage diagnostics assembly + persistence.

The Python port of ``r/2-postpro_pipeline/25-postpro_diagnostics/25-diagnostics-output.R``:

* :func:`build_postpro_diagnostics` — the clean / harmonize / standardize matched-rule summaries;
* :func:`build_last_rule_wins_overwrite_subset` — one row per final-stage row that a
  ``last_rule_wins`` update overwrote, with the overwrite events collapsed (group-by row + join);
* :func:`persist_postpro_audit` — write the per-stage audit workbooks (matched + unmatched sheets)
  and the overwrite-subset workbook.

Workbooks are written with openpyxl (the writexl/openxlsx analogue).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import polars as pl
from openpyxl import Workbook

from whep_digitize.general.config import Config
from whep_digitize.general.constants import get_pipeline_constants
from whep_digitize.general.directories import ensure_directories_exist
from whep_digitize.postpro.diagnostics.rule_summaries import (
    build_stage_rule_catalog_from_payloads,
    build_unmatched_rule_summary,
    summarize_stage_rules,
)
from whep_digitize.postpro.diagnostics.standardize_summaries import (
    build_standardize_rule_catalog,
    build_unmatched_standardize_rule_summary,
    summarize_standardize_rules,
)
from whep_digitize.postpro.utilities.output_roots import initialize_postpro_output_root
from whep_digitize.postpro.utilities.templates import load_stage_rule_payloads

_CONSTANTS = get_pipeline_constants()
_POSTPRO = _CONSTANTS.postpro
_R_TRIMWS = " \t\r\n"
_OVERWRITE_META_COLUMNS = (
    "row_id",
    "overwrite_event_count",
    "overwritten_columns",
    "overwritten_rule_files",
    "overwritten_stages",
)
_OVERWRITE_EVENT_COLUMNS = ("row_id", "column_target", "rule_file_identifier", "execution_stage")
# Standardize audit workbook column subset (R ``excel_columns``; effective before offset).
_STANDARDIZE_EXCEL_COLUMNS = (
    "affected_rows",
    "rule_file_identifier",
    "commodity_key",
    "unit_source",
    "unit_target",
    "unit_factor",
    "unit_factor_effective",
    "unit_offset",
)


@dataclass(frozen=True, slots=True)
class PostproDiagnosticsSummaries:
    """The three stage matched-rule summaries (R ``build_postpro_diagnostics`` list)."""

    clean_rule_summary: pl.DataFrame
    harmonize_rule_summary: pl.DataFrame
    standardize_rule_summary: pl.DataFrame


def build_postpro_diagnostics(
    clean_audit_dt: pl.DataFrame,
    harmonize_audit_dt: pl.DataFrame,
    standardize_audit_dt: pl.DataFrame,
) -> PostproDiagnosticsSummaries:
    """Summarize the clean / harmonize / standardize audits (R ``build_postpro_diagnostics``)."""
    return PostproDiagnosticsSummaries(
        clean_rule_summary=summarize_stage_rules(clean_audit_dt),
        harmonize_rule_summary=summarize_stage_rules(harmonize_audit_dt),
        standardize_rule_summary=summarize_standardize_rules(standardize_audit_dt),
    )


def _collapse_expr(column: str) -> pl.Expr:
    """Aggregate a group's values into a sorted, unique, ``; ``-joined string (null if empty)."""
    text = pl.col(column).cast(pl.String).str.strip_chars(_R_TRIMWS)
    kept = text.filter(text.is_not_null() & (text.str.len_chars() > 0))
    return (
        pl.when(kept.len() == 0)
        .then(pl.lit(None, dtype=pl.String))
        .otherwise(kept.unique().sort().str.join("; "))
    )


def build_last_rule_wins_overwrite_subset(
    final_stage_dt: pl.DataFrame, overwrite_events_dt: pl.DataFrame
) -> pl.DataFrame:
    """Return one row per final-stage row a ``last_rule_wins`` update overwrote.

    The Python port of R ``build_last_rule_wins_overwrite_subset``: overwrite events are grouped by
    (1-based) ``row_id``, the affected columns/files/stages collapsed, and joined to the final-stage
    row values.

    Args:
        final_stage_dt: The final post-processing frame.
        overwrite_events_dt: The overwrite events collected during rule execution.

    Returns:
        The overwrite metadata + final-stage row values (empty same-shape frame when none apply).
    """
    final = final_stage_dt.with_row_index("row_id", offset=1).with_columns(
        pl.col("row_id").cast(pl.Int64)
    )
    final_columns = [column for column in final.columns if column != "row_id"]

    if final.height == 0 or overwrite_events_dt.height == 0:
        return _empty_overwrite_subset(final, final_columns)

    events = overwrite_events_dt
    additions = [
        pl.lit(None, dtype=pl.String).alias(column)
        for column in _OVERWRITE_EVENT_COLUMNS
        if column not in events.columns
    ]
    events = events.with_columns(additions) if additions else events
    events = events.with_columns(pl.col("row_id").cast(pl.Int64, strict=False)).filter(
        pl.col("row_id").is_not_null()
        & (pl.col("row_id") >= 1)
        & (pl.col("row_id") <= final.height)
    )
    if events.height == 0:
        return _empty_overwrite_subset(final, final_columns)

    row_summary = events.group_by("row_id", maintain_order=True).agg(
        pl.len().cast(pl.Int64).alias("overwrite_event_count"),
        _collapse_expr("column_target").alias("overwritten_columns"),
        _collapse_expr("rule_file_identifier").alias("overwritten_rule_files"),
        _collapse_expr("execution_stage").alias("overwritten_stages"),
    )
    row_subset = final.filter(pl.col("row_id").is_in(row_summary.get_column("row_id").to_list()))
    return (
        row_summary.join(row_subset, on="row_id", how="left")
        .sort("row_id")
        .select(*_OVERWRITE_META_COLUMNS, *final_columns)
    )


def persist_postpro_audit(
    clean_audit_dt: pl.DataFrame,
    harmonize_audit_dt: pl.DataFrame,
    standardize_audit_dt: pl.DataFrame,
    standardize_rules_dt: pl.DataFrame,
    final_stage_dt: pl.DataFrame,
    last_rule_wins_overwrites_dt: pl.DataFrame,
    config: Config,
    *,
    standardize_matched_rule_counts_dt: pl.DataFrame | None = None,
) -> dict[str, Path]:
    """Write the per-stage audit workbooks + the overwrite-subset workbook.

    The Python port of R ``persist_postpro_audit``. Each stage workbook has ``matched_rules`` and
    ``unmatched_rules`` sheets; the overwrite workbook has a single ``last_rule_wins_overwrites``
    sheet.

    Args:
        clean_audit_dt: The clean-stage audit.
        harmonize_audit_dt: The harmonize-stage audit.
        standardize_audit_dt: The standardize-stage audit.
        standardize_rules_dt: The prepared standardize-layer rules.
        final_stage_dt: The final post-processing frame.
        last_rule_wins_overwrites_dt: The overwrite events.
        config: The resolved pipeline configuration.
        standardize_matched_rule_counts_dt: Optional standardize matched-rule counts.

    Returns:
        Mapping of workbook name → written path.
    """
    diagnostics = build_postpro_diagnostics(
        clean_audit_dt, harmonize_audit_dt, standardize_audit_dt
    )
    paths = initialize_postpro_output_root(config)
    ensure_directories_exist([paths.audit_dir, paths.diagnostics_dir])

    output_paths = {
        "clean_audit": paths.audit_dir / _POSTPRO.clean_audit_file_name,
        "harmonize_audit": paths.audit_dir / _POSTPRO.harmonize_audit_file_name,
        "standardize_audit": paths.audit_dir / _POSTPRO.standardize_audit_file_name,
        "last_rule_wins_overwrites": paths.diagnostics_dir
        / _POSTPRO.last_rule_wins_overwrites_file_name,
    }

    overwrite_subset = build_last_rule_wins_overwrite_subset(
        final_stage_dt, last_rule_wins_overwrites_dt
    )
    clean_catalog = build_stage_rule_catalog_from_payloads(
        load_stage_rule_payloads(config, "clean")
    )
    harmonize_catalog = build_stage_rule_catalog_from_payloads(
        load_stage_rule_payloads(config, "harmonize")
    )
    standardize_catalog = build_standardize_rule_catalog(standardize_rules_dt)

    clean_unmatched = build_unmatched_rule_summary(clean_catalog, diagnostics.clean_rule_summary)
    harmonize_unmatched = build_unmatched_rule_summary(
        harmonize_catalog, diagnostics.harmonize_rule_summary
    )
    standardize_unmatched = build_unmatched_standardize_rule_summary(
        standardize_catalog,
        diagnostics.standardize_rule_summary,
        standardize_matched_rule_counts_dt,
    )

    _write_sheets(
        output_paths["clean_audit"],
        {"matched_rules": diagnostics.clean_rule_summary, "unmatched_rules": clean_unmatched},
    )
    _write_sheets(
        output_paths["harmonize_audit"],
        {
            "matched_rules": diagnostics.harmonize_rule_summary,
            "unmatched_rules": harmonize_unmatched,
        },
    )
    _write_sheets(
        output_paths["standardize_audit"],
        {
            "matched_rules": _standardize_excel_subset(diagnostics.standardize_rule_summary),
            "unmatched_rules": _standardize_excel_subset(standardize_unmatched),
        },
    )
    _write_sheets(
        output_paths["last_rule_wins_overwrites"],
        {"last_rule_wins_overwrites": overwrite_subset},
    )
    return output_paths


# --------------------------------------------------------------------------- private helpers


def _empty_overwrite_subset(final: pl.DataFrame, final_columns: list[str]) -> pl.DataFrame:
    """Build the empty overwrite-subset frame (metadata columns + final-stage columns)."""
    schema: dict[str, pl.DataType | type[pl.DataType]] = {
        "row_id": pl.Int64,
        "overwrite_event_count": pl.Int64,
        "overwritten_columns": pl.String,
        "overwritten_rule_files": pl.String,
        "overwritten_stages": pl.String,
    }
    for column in final_columns:
        schema[column] = final.schema[column]
    return pl.DataFrame(schema=schema)


def _standardize_excel_subset(frame: pl.DataFrame) -> pl.DataFrame:
    """Select the standardize audit workbook column subset (R ``excel_columns``)."""
    return frame.select(_STANDARDIZE_EXCEL_COLUMNS)


def _write_sheets(path: Path, sheets: dict[str, pl.DataFrame]) -> None:
    """Write each named frame to its own worksheet (header + rows) via openpyxl."""
    workbook = Workbook()
    for index, (name, frame) in enumerate(sheets.items()):
        sheet = workbook.active if index == 0 else workbook.create_sheet(name)
        sheet.title = name
        sheet.append(list(frame.columns))
        for row in frame.iter_rows():
            sheet.append(list(row))
    ensure_directories_exist([path.parent])
    workbook.save(path)
