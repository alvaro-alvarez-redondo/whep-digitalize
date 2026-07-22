"""Postpro / audit export — ports ``r/2-postpro_pipeline/20-data_audit/20-audit-export.R``.

Writes the audit workbook via **openpyxl** (the R original used ``openxlsx``): the invalid-row
subset with each flagged cell highlighted (solid fill + bold font + thick border, from
:class:`~whep_digitize.general.constants.ErrorHighlightStyle`). Excel rows/columns are 1-based
with a one-row header offset.

Behaviors preserved from R:

* Skip workbook creation entirely when both the subset and the findings are empty.
* A ``source_row_index`` (existing ``row_index`` column, else ``1..N``) keys highlighting; when
  a ``document`` column is present the rows are stably sorted by it (nulls last).
* Only non-technical columns are written; findings whose row/column are not shown are ignored.
"""

from __future__ import annotations

from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from whep_digitize.general.config import Config
from whep_digitize.general.constants import ErrorHighlightStyle
from whep_digitize.general.directories import ensure_output_directories

_SHEET_NAME = "audit_report"
_SOURCE_ROW_INDEX = "source_row_index"
_DOCUMENT_COLUMN = "document"
_ROW_INDEX_COLUMN = "row_index"
_AUDIT_COLUMN = "audit_column"
_TECHNICAL_COLUMNS = frozenset(
    {_SOURCE_ROW_INDEX, _ROW_INDEX_COLUMN, _AUDIT_COLUMN, "audit_type", "audit_message"}
)
_EMPTY_NOTE = "No audit findings detected for this dataset."
# Excel data starts on row 2 (row 1 is the header).
_HEADER_OFFSET = 1


def export_validation_audit_report(
    audit_dt: pl.DataFrame,
    config: Config,
    findings_dt: pl.DataFrame | None,
    output_path: Path,
) -> Path | None:
    """Write the styled audit workbook, or skip when there is nothing to report.

    The Python port of R ``export_validation_audit_report``.

    Args:
        audit_dt: The invalid-row subset to write (may be empty).
        config: The pipeline configuration (supplies the error-highlight style).
        findings_dt: Findings whose ``row_index`` (matching ``audit_dt``'s ``source_row_index``)
            and ``audit_column`` drive cell highlighting; may be ``None`` (then derived from
            ``audit_dt`` if it carries those columns).
        output_path: Destination workbook path.

    Returns:
        The written path, or ``None`` when both the subset and findings are empty (no file
        created).
    """
    has_findings = findings_dt is not None and findings_dt.height > 0
    if audit_dt.height == 0 and not has_findings:
        return None

    export = _add_source_row_index(audit_dt)
    if _DOCUMENT_COLUMN in export.columns:
        export = export.sort(_DOCUMENT_COLUMN, nulls_last=True, maintain_order=True)

    cols_to_show = [column for column in export.columns if column not in _TECHNICAL_COLUMNS]
    if not cols_to_show:
        cols_to_show = [_SOURCE_ROW_INDEX]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _SHEET_NAME

    if export.height == 0:
        _write_note(worksheet)
    else:
        _write_table(worksheet, export.select(cols_to_show))
        _apply_highlights(worksheet, export, cols_to_show, findings_dt, config)

    ensure_output_directories([output_path])
    workbook.save(output_path)
    return output_path


def _add_source_row_index(audit_dt: pl.DataFrame) -> pl.DataFrame:
    """Add ``source_row_index`` — the existing ``row_index`` column, else a 1-based sequence."""
    if _ROW_INDEX_COLUMN in audit_dt.columns:
        return audit_dt.with_columns(
            pl.col(_ROW_INDEX_COLUMN).cast(pl.Int64).alias(_SOURCE_ROW_INDEX)
        )
    return audit_dt.with_row_index(_SOURCE_ROW_INDEX, offset=1).with_columns(
        pl.col(_SOURCE_ROW_INDEX).cast(pl.Int64)
    )


def _write_note(worksheet: Worksheet) -> None:
    """Write the single-cell 'no findings' note (R writes a one-column ``note`` table)."""
    worksheet.append(["note"])
    worksheet.append([_EMPTY_NOTE])


def _write_table(worksheet: Worksheet, display: pl.DataFrame) -> None:
    """Write the header row followed by every data row of ``display``."""
    worksheet.append(display.columns)
    for row in display.iter_rows():
        worksheet.append(list(row))


def _effective_findings(
    export: pl.DataFrame, findings_dt: pl.DataFrame | None
) -> pl.DataFrame | None:
    """Return the findings driving highlighting, deriving them from ``export`` when not supplied."""
    if findings_dt is not None:
        return findings_dt
    if {_ROW_INDEX_COLUMN, _AUDIT_COLUMN}.issubset(export.columns):
        return export.select(
            pl.col(_ROW_INDEX_COLUMN).cast(pl.Int64),
            pl.col(_AUDIT_COLUMN).cast(pl.String),
        ).unique(maintain_order=True)
    return None


def _apply_highlights(
    worksheet: Worksheet,
    export: pl.DataFrame,
    cols_to_show: list[str],
    findings_dt: pl.DataFrame | None,
    config: Config,
) -> None:
    """Paint the highlight style onto every flagged (shown) cell."""
    effective = _effective_findings(export, findings_dt)
    if effective is None or effective.height == 0:
        return

    findings = effective.filter(
        pl.col(_ROW_INDEX_COLUMN).is_not_null()
        & pl.col(_AUDIT_COLUMN).is_not_null()
        & (pl.col(_AUDIT_COLUMN).str.len_chars() > 0)
    )
    if findings.height == 0:
        return

    excel_row_by_source = {
        source_row_index: position + 1 + _HEADER_OFFSET
        for position, source_row_index in enumerate(export.get_column(_SOURCE_ROW_INDEX).to_list())
    }
    excel_col_by_name = {name: index + 1 for index, name in enumerate(cols_to_show)}
    fill, font, border = _build_highlight(config.export_config.error_highlight)

    for row_index, audit_column in findings.select(
        pl.col(_ROW_INDEX_COLUMN).cast(pl.Int64), pl.col(_AUDIT_COLUMN).cast(pl.String)
    ).iter_rows():
        excel_row = excel_row_by_source.get(row_index)
        excel_col = excel_col_by_name.get(audit_column)
        if excel_row is None or excel_col is None:
            continue
        cell = worksheet.cell(row=excel_row, column=excel_col)
        cell.fill = fill
        cell.font = font
        cell.border = border


def _argb(colour: str) -> str:
    """Convert a ``#RRGGBB`` hex colour to openpyxl's 8-digit ``AARRGGBB`` (opaque)."""
    return f"FF{colour.lstrip('#').upper()}"


def _build_highlight(style: ErrorHighlightStyle) -> tuple[PatternFill, Font, Border]:
    """Build the (fill, font, border) triple from the configured error-highlight style.

    Mirrors the R ``openxlsx::createStyle`` call: a solid foreground fill, a bold coloured font,
    and a thick border on all four sides (``border = "TopBottomLeftRight"``).
    """
    fill = PatternFill(fill_type="solid", fgColor=_argb(style.fg_fill))
    font = Font(bold=style.text_decoration == "bold", color=_argb(style.font_colour))
    side = Side(style=style.border_style, color=_argb(style.border_colour))
    border = Border(left=side, right=side, top=side, bottom=side)
    return fill, font, border
